import openpyxl

name_of_filds = ['Name', 'Grade', 'Age']
field_01 = ['Jack', '3', '10']
field_02 = ['Tom', '5', '12']
field_03 = ['Sam', '11', '18']


wb = openpyxl.Workbook()
print(wb.sheetnames)

wb.create_sheet(title='Первый лист', index=0)
print(wb.sheetnames)

wb.remove(wb['Sheet'])
print(wb.sheetnames)

sheet = wb['Первый лист'] # по індексу листа
print(sheet)

for row_index, row in enumerate((name_of_filds, field_01, field_02, field_03)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index+1, column=col_index+1)
        cell.value = value

wb.save('task_03.xlsx')


# wb = openpyxl.load_workbook('task_03.xlsx')
# sheets = wb.sheetnames
# print(sheets)
#
# sheet_2 = wb.active # вибираєм активний лист
# print(sheet_2)
#
# sheet = wb['Первый лист']
# print(sheet)
#
# print(sheet['A3'].value)
#
# rows = sheet.max_row # ПЕРЕВІРИЛИ СКІЛЬКИ СТРОК ЗАПОВНЕНІ
# cols = sheet.max_column # ПЕРЕВІРИЛИ СКІЛКЬИ СТОВПЧИКІВ ЗАПОВНЕНІ
#
# print(rows)
# print(cols)
#
# name_of_fields = []
# fields_values = []
#
# for i in range(1, rows+1):
#     value_row = []
#     for j in range(1, cols+1):
#         cell = sheet.cell(row=i, column=j)
#         value = str(cell.value)
#         if i == 1:
#             name_of_fields.append(value)
#         else:
#             value_row.append(value)
#     if i != 1:
#         fields_values.append(value_row)
#
# print(name_of_fields)
# print(fields_values)
#
#
